"""
Dashboard Data Publisher
------------------------------------------------------------------
Consolidates Top20 recommendations + your actual portfolio holdings into
one compact JSON file (dashboard_data.json) - small enough to load fast
on mobile, and everything the index.html needs in one place.

Run this AFTER generate_vlookup_calculator.py and generate_portfolio_tracker.py
(it reads their outputs). Already wired into daily_refresh_all.bat.
"""

import json
import os
import math
import datetime

JSON_FILE = "top20_calculator_data.json"
PORTFOLIO_FILE = "Portfolio_Tracker.xlsx"
REAL_DATA_FILE = "Complete_US_Market_Report.xlsx"
OUTPUT_FILE = "dashboard_data.json"
PUBLIC_OUTPUT_FILE = "public_data.json"


def _safe_float(v):
    try:
        if v in (None, "N/A", "", "nan"):
            return None
        f = float(v)
        # NaN != NaN is always True in Python, so an equality-based check
        # (like the one above) can never catch a real floating-point NaN -
        # explicit isnan/isinf checks are required, or a literal NaN/Infinity
        # token leaks straight into the JSON output, which browsers reject
        # as invalid (only null is valid JSON for "no value", not NaN).
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    except (TypeError, ValueError):
        return None


def compute_suggested_entry(current_price, lower_bb):
    """Same formula used in generate_vlookup_calculator.py - the raw
    screener output (Complete_US_Market_Report.xlsx) never has a
    'Suggested Entry Price' column, that's a derived value computed
    downstream. All Tickers needs to compute it here too, rather than
    reading a column that doesn't actually exist in the source file."""
    cp, lb = _safe_float(current_price), _safe_float(lower_bb)
    if cp is None:
        return None
    if lb is None:
        return round(cp * 0.995, 2)
    dist = cp - lb
    return round(min(cp * 0.995, lb + dist * 0.35), 2)


def load_all_tickers():
    """Full ticker universe with ALL the calculator's fields - not trimmed.
    The dashboard shows a compact summary by default and lets you tap a
    card to expand the rest, so the full richness of the calculator is
    actually available on mobile, just organized for a small screen."""
    if not os.path.exists(REAL_DATA_FILE):
        return []
    import pandas as pd
    df = pd.read_excel(REAL_DATA_FILE)
    tickers = []
    for _, r in df.iterrows():
        tickers.append({
            "ticker": r.get("Ticker"),
            "company": r.get("Name", r.get("Ticker")),
            "category": r.get("Category"),
            "signal": r.get("Signal"),
            "matrix_score": r.get("Matrix Score"),
            "current_price": _safe_float(r.get("Current Price")),
            "suggested_entry": compute_suggested_entry(r.get("Current Price"), r.get("Support (Lower BB)")),
            "high_52w_price": _safe_float(r.get("52W High Price")),
            "high_52w_drop_pct": r.get("52W High Drop %"),
            "ema20": _safe_float(r.get("EMA20")),
            "ema40": _safe_float(r.get("EMA40")),
            "support_lower_bb": _safe_float(r.get("Support (Lower BB)")),
            "middle_bb": _safe_float(r.get("Middle BB")),
            "resistance_upper_bb": _safe_float(r.get("Resistance (Upper BB)")),
            "rsi": _safe_float(r.get("RSI")),
            "macd_line": _safe_float(r.get("MACD Line")),
            "macd_signal": _safe_float(r.get("MACD Signal")),
            "macd_histogram": _safe_float(r.get("MACD Histogram")),
            "macd_trend": r.get("MACD Histogram Trend", "N/A"),
            "current_vol": r.get("Current Vol"),
            "vol_200sma": r.get("Vol 200SMA"),
            "volume_strength": r.get("Volume Driven Strength"),
            "rr_ratio": _safe_float(r.get("R/R Ratio")),
            "rr_note": r.get("R/R Ratio Note"),
            "intrinsic_value": _safe_float(r.get("Intrinsic (Fair) Value")),
            "margin_of_safety": r.get("Margin of Safety"),
            "analyst_target": _safe_float(r.get("Analyst Target Price")),
            "analyst_upside_pct": r.get("Analyst Upside %"),
            "put_call_ratio": _safe_float(r.get("Put/Call OI Ratio")),
            "put_wall_price": _safe_float(r.get("Whale Put Wall Price (Floor)")),
            "put_wall_oi": r.get("Put Wall Open Interest Volume"),
            "call_wall_price": _safe_float(r.get("Whale Call Wall Price (Ceiling)")),
            "call_wall_oi": r.get("Call Wall Open Interest Volume"),
            "walls_crossed": bool(r.get("Walls Crossed (Low Confidence)")),
        })
    return tickers


def load_top20():
    if not os.path.exists(JSON_FILE):
        return [], False
    with open(JSON_FILE) as f:
        data = json.load(f)
    return data.get("rows", []), data.get("is_real", False)


def load_portfolio():
    if not os.path.exists(PORTFOLIO_FILE):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(PORTFOLIO_FILE, data_only=True)
    ws = wb["My Portfolio"]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    holdings = []
    for r in range(2, ws.max_row + 1):
        ticker = ws.cell(row=r, column=headers["Ticker"]).value
        if not (ticker and isinstance(ticker, str) and len(ticker) <= 10 and " " not in ticker):
            continue  # skip the note row / blanks, same guard as generate_portfolio_tracker.py
        shares = _safe_float(ws.cell(row=r, column=headers["Shares Held"]).value) or 0
        if shares <= 0:
            continue  # not actually held
        holdings.append({
            "ticker": ticker,
            "company": ws.cell(row=r, column=headers["Company"]).value,
            "shares": shares,
            "avg_cost": _safe_float(ws.cell(row=r, column=headers["Average Cost"]).value),
            "current_price": _safe_float(ws.cell(row=r, column=headers["Current Price"]).value),
            "market_value": _safe_float(ws.cell(row=r, column=headers["Market Value"]).value),
            "gain_dollar": _safe_float(ws.cell(row=r, column=headers["Unrealized Gain/Loss ($)"]).value),
            "gain_pct": _safe_float(ws.cell(row=r, column=headers["Unrealized Gain/Loss (%)"]).value),
            "weight_pct": _safe_float(ws.cell(row=r, column=headers["Portfolio Weight (%)"]).value),
            "suggested_entry": _safe_float(ws.cell(row=r, column=headers["This Month Suggested Entry"]).value),
            "analyst_target": _safe_float(ws.cell(row=r, column=headers["Analyst Target Price"]).value),
            "macd_trend": ws.cell(row=r, column=headers["MACD Histogram Trend"]).value,
        })
    return holdings


def main():
    top20_rows, is_real = load_top20()
    holdings = load_portfolio()
    all_tickers = load_all_tickers()
    held_tickers = {h["ticker"] for h in holdings}

    top20 = []
    for r in top20_rows:
        top20.append({
            "ticker": r["Ticker"],
            "company": r.get("Name", r["Ticker"]),
            "signal": r.get("Signal"),
            "matrix_score": r.get("Matrix Score"),
            "current_price": _safe_float(r.get("Current Price")),
            "suggested_entry": _safe_float(r.get("This Month Entry Price")),
            "analyst_target": _safe_float(r.get("Analyst Target Price")),
            "macd_trend": r.get("MACD Histogram Trend", "N/A"),
            "shares_needed": _safe_float(r.get("Shares Needed (This Month)")),
            "capital_required": _safe_float(r.get("Capital Required (This Month)")),
            "already_held": r["Ticker"] in held_tickers,
        })

    total_value = sum(h["market_value"] for h in holdings if h["market_value"])
    total_cost = sum((h["avg_cost"] or 0) * h["shares"] for h in holdings)
    total_gain = total_value - total_cost if total_value else 0

    payload = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "is_real_data": is_real,
        "portfolio": {
            "total_value": round(total_value, 2),
            "total_cost": round(total_cost, 2),
            "total_gain": round(total_gain, 2),
            "total_gain_pct": round(total_gain / total_cost, 4) if total_cost else None,
            "holdings": holdings,
        },
        "top20": top20,
        "all_tickers": all_tickers,
    }

    # Safety net: recursively replace any NaN/Infinity anywhere in the
    # payload with null. This is a backstop, not the primary fix - the
    # primary fix is _safe_float() actually catching these at the source -
    # but this guarantees the file can NEVER contain invalid JSON even if a
    # future field gets added without going through _safe_float.
    def sanitize(obj):
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        if isinstance(obj, dict):
            return {k: sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [sanitize(v) for v in obj]
        return obj

    payload = sanitize(payload)

    with open(OUTPUT_FILE, "w") as f:
        json.dump(payload, f, indent=2)

    print(f"[SUCCESS] Wrote {OUTPUT_FILE} - {len(holdings)} holdings, {len(top20)} recommended picks, "
          f"{len(all_tickers)} total tickers")

    # --- Public version: genuinely no portfolio data, not just hidden in
    # the UI. Even 'already_held' is stripped from top20, since it would
    # otherwise leak which tickers you hold indirectly. This file is safe
    # to put behind a Cloudflare Access bypass rule and share with anyone -
    # there is nothing personal in it to find, even by viewing raw JSON. ---
    public_top20 = [{k: v for k, v in r.items() if k != "already_held"} for r in payload["top20"]]
    public_payload = {
        "generated_at": payload["generated_at"],
        "is_real_data": payload["is_real_data"],
        "top20": public_top20,
        "all_tickers": payload["all_tickers"],
    }
    with open(PUBLIC_OUTPUT_FILE, "w") as f:
        json.dump(public_payload, f, indent=2)
    print(f"[SUCCESS] Wrote {PUBLIC_OUTPUT_FILE} (no portfolio data) - safe to share publicly")


if __name__ == "__main__":
    main()
